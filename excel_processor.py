import pandas as pd
import tempfile
from typing import List, Dict
from text_unifier import TextUnifier


class ExcelProcessor:
    def __init__(self, api_key: str = ""):
        """
        ExcelProcessorの初期化
        """
        self.api_key = api_key
        self.text_unifier = TextUnifier(api_key)
        self.sheet_names = {
            'extraction': '抽出テキスト',
            'original': '誤った転記',
            'corrected': '正しい転記',
            'differences': '差異記録'
        }

    # ==========================
    # DataFrame I/O helpers
    # ==========================
    def read_excel(self, file_path):
        """
        Excelファイルを読み込み、すべてのシートのデータを返す
        """
        try:
            xls = pd.ExcelFile(file_path)
            sheets_data = {}
            for sheet_name in xls.sheet_names:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                df = df.fillna('')
                sheets_data[sheet_name] = df
            return sheets_data
        except Exception as e:
            raise Exception(f"Excelファイルの読み込みエラー: {str(e)}")

    def _prepare_extraction_sheet(self, extraction_data):
        if not extraction_data or len(extraction_data) == 0:
            return pd.DataFrame()
        first_sheet = list(extraction_data.values())[0]
        return first_sheet.copy()

    def _prepare_original_sheet(self, submission_data):
        if not submission_data or len(submission_data) == 0:
            return pd.DataFrame()
        first_sheet = list(submission_data.values())[0]
        return first_sheet.copy()

    def _prepare_corrected_sheet(self, unified_data):
        if unified_data is None or (isinstance(unified_data, pd.DataFrame) and unified_data.empty):
            return pd.DataFrame()
        return unified_data.copy()

    # ==========================
    # Text helpers
    # ==========================
    def normalize_text(self, text):
        if not isinstance(text, str):
            return str(text)
        text = text.translate(str.maketrans(
            'ＡＢＣＤＥＦＧＨＩＪＫＬＭＮＯＰＱＲＳＴＵＶＷＸＹＺ'
            'ａｂｃｄｅｆｇｈｉｊｋｌｍｎｏｐｑｒｓｔｕｖｗｘｙｚ'
            '０１２３４５６７８９',
            'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
            'abcdefghijklmnopqrstuvwxyz'
            '0123456789'
        ))
        text = text.translate(str.maketrans('（）［］｛｝', '()[]{}'))
        return text.strip()

    def find_similar_texts(self, target_text, text_list, threshold=0.8):
        import difflib
        target_normalized = self.normalize_text(target_text)
        similar_texts = []
        for text_info in text_list:
            text_normalized = self.normalize_text(text_info['text'])
            similarity = difflib.SequenceMatcher(
                None,
                target_normalized.lower(),
                text_normalized.lower()
            ).ratio()
            if similarity >= threshold:
                similar_texts.append({
                    **text_info,
                    'similarity': similarity,
                    'normalized_text': text_normalized
                })
        return sorted(similar_texts, key=lambda x: x['similarity'], reverse=True)

    def compare_texts(self, text1, text2):
        import difflib
        norm_text1 = self.normalize_text(text1)
        norm_text2 = self.normalize_text(text2)
        if norm_text1 == norm_text2:
            return None
        diff = list(difflib.unified_diff(
            norm_text1.splitlines(keepends=True),
            norm_text2.splitlines(keepends=True),
            fromfile='元のテキスト',
            tofile='修正後のテキスト',
            lineterm=''
        ))
        return {
            'original': text1,
            'corrected': text2,
            'normalized_original': norm_text1,
            'normalized_corrected': norm_text2,
            'diff': ''.join(diff) if diff else None
        }

    # ==========================
    # Output (format-preserving)
    # ==========================
    def create_output_file_preserve_format(self, extraction_file_path: str, submission_file_path: str,
                                           unified_file_path: str, differences: List[Dict]) -> str:
        """
        元のフォーマットを保持した出力用Excelファイルを作成
        """
        try:
            import openpyxl
            import tempfile

            output_path = tempfile.mktemp(suffix='.xlsx')
            wb = openpyxl.Workbook()
            # 既定シート削除
            wb.remove(wb.active)

            # 1〜3枚目をコピー
            self._copy_excel_sheets(extraction_file_path, wb, self.sheet_names['extraction'])
            self._copy_excel_sheets(submission_file_path, wb, self.sheet_names['original'])
            self._copy_excel_sheets(unified_file_path, wb, self.sheet_names['corrected'])

            # 4枚目: 差異記録
            if differences and len(differences) > 0:
                self._add_differences_sheet(wb, differences)
            else:
                diff_sheet = wb.create_sheet(title=self.sheet_names['differences'])
                diff_sheet['A1'] = "差異はありませんでした"
                diff_sheet['A1'].font = openpyxl.styles.Font(bold=True)

            # === レイアウト同期 ===
            # ・C〜G の結合をテンプレ（submission）通りに復元
            # ・K〜AO の列幅をテンプレと同じに
            # ・B列は折返し＋テンプレ幅を下限に、文字数で広げる
            self.__apply_layout_from_template(
                target_sheet=wb[self.sheet_names['corrected']],
                template_file_path=submission_file_path,
                sync_columns=("K", "AO"),
                restore_merges=True,
                adjust_col_B_by_text=True
            )

            wb.save(output_path)
            wb.close()
            return output_path

        except Exception as e:
            raise Exception(f"フォーマット保持出力ファイル作成エラー: {str(e)}")

    def process_files(self, extraction_file_path: str, submission_file_path: str) -> tuple:
        """
        抽出ファイルと転記ファイルを処理し、統一されたファイルと差異を返す
        """
        try:
            extraction_data = self.read_excel(extraction_file_path)
            submission_data = self.read_excel(submission_file_path)

            # 統一処理（テンプレに合わせた内容で別xlsxが出る想定）
            unified_file_path, differences = self.text_unifier.unify_texts_preserve_format(
                extraction_data, submission_data, submission_file_path
            )

            # 4シートの最終ブックを作成
            output_path = self.create_output_file_preserve_format(
                extraction_file_path, submission_file_path, unified_file_path, differences
            )

            # 罫線を黒に統一
            self._force_black_borders(output_path)
            return output_path, differences

        except Exception as e:
            raise Exception(f"ファイル処理エラー: {str(e)}")

    # ==========================
    # Low-level copy helpers
    # ==========================
    def _copy_excel_sheets(self, source_file_path: str, target_workbook, new_sheet_name: str):
        """
        ソースファイルの最初のシートを対象ワークブックにコピー（値/書式/結合/列幅/行高）
        """
        try:
            import openpyxl
            source_wb = openpyxl.load_workbook(source_file_path)
            src = source_wb.active
            dst = target_workbook.create_sheet(title=new_sheet_name)

            # 値 & スタイル
            for row in src.iter_rows():
                for cell in row:
                    d = dst.cell(row=cell.row, column=cell.column, value=cell.value)

                    # font
                    if cell.font:
                        d.font = openpyxl.styles.Font(
                            name=cell.font.name,
                            size=cell.font.size,
                            bold=cell.font.bold,
                            italic=cell.font.italic,
                            underline=cell.font.underline,
                            strike=cell.font.strike,
                            color=cell.font.color
                        )
                    # fill
                    if cell.fill:
                        d.fill = openpyxl.styles.PatternFill(
                            fill_type=cell.fill.fill_type,
                            start_color=cell.fill.start_color,
                            end_color=cell.fill.end_color
                        )
                    # border
                    if cell.border:
                        d.border = openpyxl.styles.Border(
                            left=cell.border.left,
                            right=cell.border.right,
                            top=cell.border.top,
                            bottom=cell.border.bottom,
                            diagonal=cell.border.diagonal,
                            vertical=cell.border.vertical,
                            horizontal=cell.border.horizontal,
                        )
                    # alignment
                    if cell.alignment:
                        d.alignment = openpyxl.styles.Alignment(
                            horizontal=cell.alignment.horizontal,
                            vertical=cell.alignment.vertical,
                            wrap_text=cell.alignment.wrap_text,
                            shrink_to_fit=cell.alignment.shrink_to_fit,
                            text_rotation=cell.alignment.text_rotation,
                            indent=cell.alignment.indent
                        )

            # 結合セルをコピー
            if src.merged_cells.ranges:
                for r in list(src.merged_cells.ranges):
                    dst.merge_cells(str(r))

            # 列幅をコピー（値があるものはそのまま）
            for col_letter, dim in src.column_dimensions.items():
                if dim.width is not None:
                    dst.column_dimensions[col_letter].width = dim.width

            # 行高をコピー（値があるものはそのまま）
            for r_idx, dim in src.row_dimensions.items():
                if dim.height is not None:
                    dst.row_dimensions[r_idx].height = dim.height

            source_wb.close()

        except Exception as e:
            raise Exception(f"シートコピーエラー: {str(e)}")

    def _add_differences_sheet(self, workbook, differences: List[Dict]):
        """
        差異記録シートを追加
        """
        try:
            import openpyxl
            diff_sheet = workbook.create_sheet(title=self.sheet_names['differences'])

            headers = ['シート', '列', '行', 'セル', '元のテキスト', '修正後テキスト', '修正理由', '信頼度']
            for col, header in enumerate(headers, 1):
                cell = diff_sheet.cell(row=1, column=col)
                cell.value = header
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(fill_type="solid", start_color="CCCCCC")

            for row, diff in enumerate(differences, 2):
                diff_sheet.cell(row=row, column=1).value = diff.get('シート', '')
                diff_sheet.cell(row=row, column=2).value = diff.get('列', '')
                diff_sheet.cell(row=row, column=3).value = diff.get('行', '')
                diff_sheet.cell(row=row, column=4).value = diff.get('セル', '')
                diff_sheet.cell(row=row, column=5).value = diff.get('元のテキスト', '')
                diff_sheet.cell(row=row, column=6).value = diff.get('修正後テキスト', '')
                diff_sheet.cell(row=row, column=7).value = diff.get('修正理由', '')
                diff_sheet.cell(row=row, column=8).value = diff.get('信頼度', '')

        except Exception as e:
            raise Exception(f"差異シート追加エラー: {str(e)}")

    def _force_black_borders(self, xlsx_path: str):
        """
        全シートの罫線色を黒(ARGB:FF000000)に統一
        """
        import openpyxl
        from openpyxl.styles import Border, Side

        wb = openpyxl.load_workbook(xlsx_path)
        black = "FF000000"
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    b = cell.border
                    if not b:
                        continue

                    def black_side(side):
                        if side is None:
                            return None
                        return Side(style=side.style, color=black)

                    cell.border = Border(
                        left=black_side(b.left),
                        right=black_side(b.right),
                        top=black_side(b.top),
                        bottom=black_side(b.bottom),
                        diagonal=black_side(b.diagonal),
                        vertical=black_side(b.vertical),
                        horizontal=black_side(b.horizontal),
                    )
        wb.save(xlsx_path)
        wb.close()

    # ==========================
    # Layout sync helper
    # ==========================
    def __apply_layout_from_template(
        self,
        target_sheet,
        template_file_path: str,
        sync_columns: tuple = ("K", "AO"),
        restore_merges: bool = True,
        adjust_col_B_by_text: bool = True,
        max_col_B_width: int = 80  # 文字数上限（Excel幅相当）
    ):
        """
        テンプレ（提出ファイル=template_file_path の1枚目）からレイアウトを同期する。
        - restore_merges=True: テンプレと同じ結合セルを復元（C〜Gの結合を含む）
        - sync_columns=(start,end): 指定列範囲の列幅をテンプレ値に統一（K〜AO）
        - adjust_col_B_by_text=True: B列は wrap_text + テンプレ幅を下限に、文字数に応じて広げる
        """
        import openpyxl
        from openpyxl.utils import column_index_from_string, get_column_letter
        from openpyxl.styles import Alignment

        tpl_wb = openpyxl.load_workbook(template_file_path)
        tpl = tpl_wb.active

        # 1) 結合セルをテンプレ通りに
        if restore_merges:
            # 既存結合を一旦解除
            if target_sheet.merged_cells.ranges:
                for r in list(target_sheet.merged_cells.ranges):
                    target_sheet.unmerge_cells(str(r))
            # テンプレの結合を適用
            for r in list(tpl.merged_cells.ranges):
                target_sheet.merge_cells(str(r))

        # 2) K〜AO の列幅をテンプレと同じに
        if sync_columns:
            s_col, e_col = sync_columns
            s_idx = column_index_from_string(s_col)
            e_idx = column_index_from_string(e_col)
            for i in range(s_idx, e_idx + 1):
                letter = get_column_letter(i)
                tpl_dim = tpl.column_dimensions.get(letter)
                if tpl_dim and tpl_dim.width is not None:
                    target_sheet.column_dimensions[letter].width = tpl_dim.width

        # 3) B列：折返し＋テンプレ幅を下限に文字数で広げる
        if adjust_col_B_by_text:
            # B列のテンプレ幅を取得（無ければ12を既定値に）
            tpl_width = tpl.column_dimensions.get("B").width if tpl.column_dimensions.get("B") else 12

            # 全セルを折返し＆上詰めに
            max_len = 0
            for c in target_sheet["B"]:
                val = "" if c.value is None else str(c.value)
                # 改行があれば行ごとに最大長を取る
                if "\n" in val:
                    local_max = max(len(line) for line in val.splitlines())
                else:
                    local_max = len(val)
                if local_max > max_len:
                    max_len = local_max

                c.alignment = Alignment(
                    horizontal=c.alignment.horizontal if c.alignment else None,
                    vertical="top",
                    wrap_text=True
                )

            # Excelの列幅は「標準フォントの字数」換算のため近似で十分
            # 文字数 + 2（余白）をテンプレ幅との max で採用（上限 max_col_B_width）
            target_width = max(tpl_width or 12, min(max_len + 2, max_col_B_width))
            target_sheet.column_dimensions["B"].width = target_width

        tpl_wb.close()
