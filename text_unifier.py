import google.generativeai as genai
import pandas as pd
import json
from typing import Dict, List, Tuple, Any
import difflib

class TextUnifier:
    def __init__(self, api_key: str):
        """TextUnifierの初期化"""
        self.api_key = api_key
        self.model = None
        self._initialize_model()

        # --- 強制置換ルール（まず絶対に直したい表記をここで定義）---
        # 必要に応じて ("置換前", "置換後") を追加してください。
        self.force_map: List[Tuple[str, str]] = [
            ("量産開始時", "作業開始時"),
        ]
    
    def _initialize_model(self):
        """Gemini APIモデルの初期化"""
        try:
            genai.configure(api_key=self.api_key)
            self.model = genai.GenerativeModel('gemini-1.5-flash')
        except Exception:
            # ここで例外を投げず、AIなしでも動けるようにする
            self.model = None

    # ===== DataFrame 経由の統一（任意で使用） =====
    def unify_texts(self, extraction_data: Dict, submission_data: Dict) -> Tuple[pd.DataFrame, List[Dict]]:
        """テキスト統一処理のメイン関数（DataFrame版）"""
        try:
            extraction_texts = self._extract_text_from_data(extraction_data)
            submission_df = list(submission_data.values())[0].copy()
            unified_df, differences = self._process_unification(submission_df, extraction_texts)
            return unified_df, differences
        except Exception as e:
            raise Exception(f"テキスト統一処理エラー: {str(e)}")
    
    def _extract_text_from_data(self, data: Dict) -> List[str]:
        """抽出データからユニークなテキストリストを作成"""
        texts = []
        for _, df in data.items():
            for col in df.columns:
                for value in df[col]:
                    if isinstance(value, str) and value.strip():
                        texts.append(value.strip())
        return list(set(texts))
    
    def _process_unification(self, submission_df: pd.DataFrame, extraction_texts: List[str]) -> Tuple[pd.DataFrame, List[Dict]]:
        """DataFrameベースの統一実行"""
        unified_df = submission_df.copy()
        differences = []
        for col in submission_df.columns:
            for idx, value in enumerate(submission_df[col]):
                if isinstance(value, str) and value.strip():
                    unified_text, info = self._unify_single_text(value.strip(), extraction_texts)
                    if info:
                        unified_df.iloc[idx, submission_df.columns.get_loc(col)] = unified_text
                        differences.append({
                            'シート': 'Sheet1',
                            '列': col,
                            '行': idx + 2,
                            '元のテキスト': value,
                            '修正後テキスト': unified_text,
                            '修正理由': info['reason'],
                            '信頼度': info['confidence']
                        })
        return unified_df, differences

    # --- 強制置換の適用 ---
    def _apply_force_map(self, text: str) -> Tuple[str, Dict]:
        """
        強制置換ルールを適用する。
        置換が発生したら (置換後テキスト, 修正情報) を返す。なければ (元テキスト, None)。
        """
        new_text = text
        applied = []
        for src, dst in self.force_map:
            if src in new_text:
                new_text = new_text.replace(src, dst)
                applied.append((src, dst))
        if applied:
            pairs = ", ".join([f"{s}→{d}" for s, d in applied])
            return new_text, {
                "reason": f"強制置換ルール適用（{pairs}）",
                "confidence": 1.0
            }
        return text, None
    
    # ===== セル値の統一ロジック =====
    def _unify_single_text(self, target_text: str, extraction_texts: List[str]) -> Tuple[str, Dict]:
        """
        単一セルの統一処理：
        0) 強制置換ルールを最初に適用（ここで確実に直す）
        1) 抽出側に完全一致があればそのまま
        2) 近い候補を探索 → 【AI不要の安全判定】で即修正 or AIで最終確認
        """
        # 0) まず強制置換を適用
        forced_text, forced_info = self._apply_force_map(target_text)
        if forced_info:
            return forced_text, forced_info

        # 1) 完全一致
        if target_text in extraction_texts:
            return target_text, None
        
        # 2) 類似候補検索
        matching = self._find_semantically_matching_extraction(target_text, extraction_texts)
        if not matching:
            return target_text, None
        
        candidate = matching['text']
        sim = matching['similarity']
        t_norm = self._deep_normalize_text(target_text)
        c_norm = self._deep_normalize_text(candidate)

        # ---- AI不要の安全即決ルート（必ず修正）----
        if t_norm == c_norm:
            # 文字種/表記ゆれだけ
            unified = self._apply_partial_replacement(target_text, candidate)
            return unified, {
                'reason': f'正規化一致（表記ゆれ統一） 類似度:{sim:.2f}',
                'confidence': 1.0
            }
        if self._is_character_type_change_only(target_text, candidate):
            unified = self._apply_partial_replacement(target_text, candidate)
            return unified, {
                'reason': f'文字種のみの差異 統一 類似度:{sim:.2f}',
                'confidence': 0.95
            }
        if self._is_minor_modification(target_text, candidate):
            unified = self._apply_partial_replacement(target_text, candidate)
            return unified, {
                'reason': f'助詞などの軽微差異 統一 類似度:{sim:.2f}',
                'confidence': 0.92
            }
        if sim >= 0.92:
            unified = self._apply_partial_replacement(target_text, candidate)
            return unified, {
                'reason': f'高類似度（>=0.92） 統一',
                'confidence': min(sim + 0.05, 1.0)
            }

        # ---- ここからAI（あれば）で確認 ----
        ai_ok, ai_reason, ai_conf = False, "AI未使用/失敗", 0.0
        if self.model is not None:
            ai = self._verify_with_ai(target_text, candidate)
            ai_ok = bool(ai.get('should_correct', False))
            ai_reason = ai.get('reason', ai_reason)
            ai_conf = ai.get('confidence', 0.0)

        if ai_ok:
            unified = self._apply_partial_replacement(target_text, candidate)
            return unified, {
                'reason': f'AI判定：{ai_reason} 類似度:{sim:.2f}',
                'confidence': max(ai_conf, min(sim + 0.1, 1.0))
            }

        # ---- AI不許可/AI不在時のフォールバック ----
        if self._is_safe_to_convert(target_text, candidate, sim) and sim >= 0.75:
            unified = self._apply_partial_replacement(target_text, candidate)
            return unified, {
                'reason': f'フォールバック：高類似かつ安全（>=0.75） 類似度:{sim:.2f}',
                'confidence': min(sim + 0.05, 0.9)
            }

        # 修正なし
        return target_text, None
    
    def _apply_partial_replacement(self, target_text: str, extraction_text: str) -> str:
        """部分一致時：括弧内等の付帯情報はできる限り保持して置換"""
        import re
        t_norm = self._deep_normalize_text(target_text)
        e_norm = self._deep_normalize_text(extraction_text)
        if e_norm in t_norm:
            preserved = []
            for pat in [r'（[^）]*）', r'\([^)]*\)', r'【[^】]*】', r'\[[^\]]*\]']:
                preserved.extend(re.findall(pat, target_text))
            result = extraction_text
            for part in preserved:
                inner = re.sub(r'[（）\(\)]', '', part)
                if part not in result and inner not in result:
                    result += part
            return result
        return extraction_text
    
    def _find_semantically_matching_extraction(self, target_text: str, extraction_texts: List[str]) -> Dict:
        """抽出テキスト中から意味的に最も近い候補を返す（部分一致含む）"""
        target_norm = self._deep_normalize_text(target_text)
        best, best_sim = None, 0.0
        for ext in extraction_texts:
            ext_norm = self._deep_normalize_text(ext)
            sim = self._calculate_semantic_similarity(target_norm, ext_norm)
            partial_sim = 0.0
            if len(ext_norm) < len(target_norm) and ext_norm in target_norm:
                partial_sim = 0.95
            elif len(target_norm) < len(ext_norm) and target_norm in ext_norm:
                partial_sim = 0.95
            final_sim = max(sim, partial_sim)
            if not self._is_safe_to_convert(target_text, ext, final_sim):
                continue
            threshold = 0.5  # 少し緩め
            if self._contains_halfwidth_katakana(ext):
                threshold = 0.45
            if final_sim > best_sim and final_sim > threshold:
                best_sim = final_sim
                best = {
                    'text': ext,
                    'similarity': final_sim,
                    'confidence': min(final_sim + 0.1, 1.0),
                    'match_type': 'partial' if final_sim == partial_sim else 'complete'
                }
        if best and target_norm == self._deep_normalize_text(best['text']):
            best['similarity'] = 1.0
            best['confidence'] = 1.0
            best['match_type'] = 'exact'
        return best
    
    def _is_safe_to_convert(self, target_text: str, extraction_text: str, similarity: float) -> bool:
        """誤変換防止の安全判定（少し緩和）"""
        if self._is_character_type_change_only(target_text, extraction_text):
            return True
        if similarity >= 0.85:
            return True
        if self._is_minor_modification(target_text, extraction_text):
            return True
        t_norm = self._deep_normalize_text(target_text)
        e_norm = self._deep_normalize_text(extraction_text)
        if (e_norm in t_norm or t_norm in e_norm) and similarity >= 0.65:
            return True
        if self._has_different_stems(target_text, extraction_text):
            return False
        if self._contains_halfwidth_katakana(extraction_text) and similarity >= 0.55:
            return True
        return False
    
    def _is_character_type_change_only(self, a: str, b: str) -> bool:
        return self._deep_normalize_text(a) == self._deep_normalize_text(b)
    
    def _has_different_stems(self, a: str, b: str) -> bool:
        import re
        sa = re.findall(r'[一-龯]+', a)
        sb = re.findall(r'[一-龯]+', b)
        if sa and sb:
            return max(sa, key=len) != max(sb, key=len)
        return False
    
    def _is_minor_modification(self, a: str, b: str) -> bool:
        import re
        a2 = re.sub(r'[をのはがにで]', '', a)
        b2 = re.sub(r'[をのはがにで]', '', b)
        return self._deep_normalize_text(a2) == self._deep_normalize_text(b2)
    
    def _contains_halfwidth_katakana(self, text: str) -> bool:
        import re
        return bool(re.search(r'[ｱ-ﾝｧ-ｮﾞﾟ]', text))
    
    def _deep_normalize_text(self, text: str) -> str:
        if not isinstance(text, str):
            return str(text)
        import unicodedata, re
        text = unicodedata.normalize('NFKC', text)
        text = self._normalize_katakana_variants(text)
        text = self._katakana_to_hiragana(text)
        text = self._convert_kanji_numbers(text)
        text = self._remove_optional_particles(text)
        text = re.sub(r'[　\s]+', '', text)
        text = re.sub(r'[・･]', '', text)
        text = re.sub(r'[ー－−‐]', '', text)
        return text.lower().strip()
    
    def _normalize_katakana_variants(self, text: str) -> str:
        result, i = "", 0
        while i < len(text):
            ch = text[i]
            nxt = text[i+1] if i+1 < len(text) else None
            if 'ｱ' <= ch <= 'ﾝ' or ch in 'ｧｨｩｪｫｯｬｭｮ':
                if nxt == 'ﾞ':
                    result += {'ｶ':'ガ','ｷ':'ギ','ｸ':'グ','ｹ':'ゲ','ｺ':'ゴ','ｻ':'ザ','ｼ':'ジ','ｽ':'ズ','ｾ':'ゼ','ｿ':'ゾ','ﾀ':'ダ','ﾁ':'ヂ','ﾂ':'ヅ','ﾃ':'デ','ﾄ':'ド','ﾊ':'バ','ﾋ':'ビ','ﾌ':'ブ','ﾍ':'ベ','ﾎ':'ボ','ｳ':'ヴ'}.get(ch, ch+nxt)
                    i += 1
                elif nxt == 'ﾟ':
                    result += {'ﾊ':'パ','ﾋ':'ピ','ﾌ':'プ','ﾍ':'ペ','ﾎ':'ポ'}.get(ch, ch+nxt)
                    i += 1
                else:
                    result += {'ｱ':'ア','ｲ':'イ','ｳ':'ウ','ｴ':'エ','ｵ':'オ','ｶ':'カ','ｷ':'キ','ｸ':'ク','ｹ':'ケ','ｺ':'コ','ｻ':'サ','ｼ':'シ','ｽ':'ス','ｾ':'セ','ｿ':'ソ','ﾀ':'タ','ﾁ':'チ','ﾂ':'ツ','ﾃ':'テ','ﾄ':'ト','ﾅ':'ナ','ﾆ':'ニ','ﾇ':'ヌ','ﾈ':'ネ','ﾉ':'ノ','ﾊ':'ハ','ﾋ':'ヒ','ﾌ':'フ','ﾍ':'ヘ','ﾎ':'ホ','ﾏ':'マ','ﾐ':'ミ','ﾑ':'ム','ﾒ':'メ','ﾓ':'モ','ﾔ':'ヤ','ﾕ':'ユ','ﾖ':'ヨ','ﾗ':'ラ','ﾘ':'リ','ﾙ':'ル','ﾚ':'レ','ﾛ':'ロ','ﾜ':'ワ','ﾝ':'ン','ｧ':'ァ','ｨ':'ィ','ｩ':'ゥ','ｪ':'ェ','ｫ':'ォ','ｯ':'ッ','ｬ':'ャ','ｭ':'ュ','ｮ':'ョ'}.get(ch, ch)
            else:
                result += ch
            i += 1
        return result
    
    def _katakana_to_hiragana(self, text: str) -> str:
        return "".join(chr(ord(c) - ord('ア') + ord('あ')) if 'ァ' <= c <= 'ヶ' else c for c in text)
    
    def _convert_kanji_numbers(self, text: str) -> str:
        for k, v in {'零':'0','〇':'0','一':'1','二':'2','三':'3','四':'4','五':'5','六':'6','七':'7','八':'8','九':'9','十':'10'}.items():
            text = text.replace(k, v)
        return text
    
    def _remove_optional_particles(self, text: str) -> str:
        import re
        return re.sub(r'[をのはがにで]', '', text)
    
    def _calculate_semantic_similarity(self, t1: str, t2: str) -> float:
        if not t1 or not t2:
            return 0.0
        s = difflib.SequenceMatcher(None, t1, t2).ratio()
        u = set(t1) | set(t2)
        c = len(set(t1) & set(t2)) / len(u) if u else 0.0
        l = min(len(t1), len(t2)) / max(len(t1), len(t2)) if max(len(t1), len(t2)) > 0 else 1.0
        return s*0.5 + c*0.3 + l*0.2
    
    def _verify_with_ai(self, original: str, candidate: str) -> Dict:
        """AIによる修正妥当性確認（“修正寄り”のプロンプト）"""
        if self.model is None:
            return {"should_correct": False, "reason": "AI未初期化", "confidence": 0.0}
        try:
            prompt = f"""
以下の2つのテキストを比較し、抽出テキスト（正しい形式）に統一すべきか判定してください。
意味が明確に異なる専門用語・概念の場合のみ false。それ以外は true。

現在のテキスト: "{original}"
抽出テキスト: "{candidate}"

出力はJSONのみ:
{{"should_correct": true/false, "reason": "理由", "confidence": 0.0～1.0}}
"""
            response = self.model.generate_content(
                prompt,
                generation_config=genai.types.GenerationConfig(
                    temperature=0.0, max_output_tokens=150, candidate_count=1
                )
            )
            import re
            m = re.search(r'\{[^}]+\}', (response.text or ""))
            if m:
                return json.loads(m.group())
            return {"should_correct": False, "reason": "AI応答解析失敗", "confidence": 0.0}
        except Exception as e:
            return {"should_correct": False, "reason": f"APIエラー: {str(e)}", "confidence": 0.0}
    
    # ===== Excelフォーマット保持版 =====
    def unify_texts_preserve_format(self, extraction_data: Dict, submission_data: Dict, submission_file_path: str) -> Tuple[str, List[Dict]]:
        """
        元のExcelレイアウトを保ったまま、セル文字列だけ統一して保存。
        戻り値: (修正後ファイルパス, 差異リスト)
        """
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            import tempfile, shutil

            extraction_texts = self._extract_text_from_data(extraction_data)

            output_path = tempfile.mktemp(suffix='.xlsx')
            shutil.copy2(submission_file_path, output_path)

            wb = openpyxl.load_workbook(output_path)
            differences: List[Dict[str, Any]] = []

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows():
                    for cell in row:
                        v = cell.value
                        if isinstance(v, str):
                            original = v.strip()
                            if not original:
                                continue
                            unified, info = self._unify_single_text(original, extraction_texts)
                            if info:
                                cell.value = unified
                                differences.append({
                                    'シート': sheet_name,
                                    '列': get_column_letter(cell.column),
                                    '行': cell.row,
                                    'セル': f"{get_column_letter(cell.column)}{cell.row}",
                                    '元のテキスト': original,
                                    '修正後テキスト': unified,
                                    '修正理由': info['reason'],
                                    '信頼度': info['confidence']
                                })
            wb.save(output_path)
            wb.close()
            return output_path, differences

        except Exception as e:
            raise Exception(f"フォーマット保持統一処理エラー: {str(e)}")
    
    # ===== サマリー（任意） =====
    def create_correction_summary(self, differences: List[Dict]) -> str:
        if not differences:
            return "修正対象のテキストはありませんでした。"
        summary = f"合計 {len(differences)} 件の修正を実行しました。\n\n"
        reason_counts: Dict[str, int] = {}
        for d in differences:
            reason = str(d.get('修正理由', '')).split(',')[0]
            reason_counts[reason] = reason_counts.get(reason, 0) + 1
        summary += "修正理由別の件数:\n"
        for r, c in reason_counts.items():
            summary += f"- {r}: {c}件\n"
        return summary
